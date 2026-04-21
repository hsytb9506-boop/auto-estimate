from fastapi import FastAPI, Depends, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime
import io, csv

from database import get_db, engine
import models, schemas
from auth import (
    hash_password, verify_password, create_access_token,
    get_current_user, require_user, require_admin
)

models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="Auto Estimate System", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://127.0.0.1:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─────────────────────────────────────────
# 초기 DB 시드 (최초 실행 시 1회)
# ─────────────────────────────────────────
@app.on_event("startup")
def startup_seed():
    db = next(get_db())
    try:
        if db.query(models.Domain).count() == 0:
            from seed import run_seed
            run_seed(db)
    finally:
        db.close()

# ─────────────────────────────────────────
# AUTH
# ─────────────────────────────────────────
@app.post("/api/auth/login", response_model=schemas.TokenResponse)
def login(data: schemas.LoginRequest, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.username == data.username).first()
    if not user or not verify_password(data.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="아이디 또는 비밀번호가 틀렸습니다")
    token = create_access_token({"sub": user.username, "is_admin": user.is_admin})
    return {"access_token": token, "token_type": "bearer",
            "is_admin": user.is_admin, "username": user.username}

@app.post("/api/auth/register")
def register(data: schemas.LoginRequest, db: Session = Depends(get_db),
             admin: models.User = Depends(require_admin)):
    if db.query(models.User).filter(models.User.username == data.username).first():
        raise HTTPException(status_code=400, detail="이미 존재하는 사용자입니다")
    user = models.User(username=data.username, hashed_password=hash_password(data.password))
    db.add(user); db.commit()
    return {"message": "사용자 생성 완료"}

@app.get("/api/auth/me")
def me(user: models.User = Depends(require_user)):
    return {"username": user.username, "is_admin": user.is_admin}

# ─────────────────────────────────────────
# DOMAINS
# ─────────────────────────────────────────
@app.get("/api/domains", response_model=List[schemas.DomainOut])
def get_domains(db: Session = Depends(get_db)):
    return db.query(models.Domain).all()

# ─────────────────────────────────────────
# CATEGORIES
# ─────────────────────────────────────────
@app.get("/api/categories", response_model=List[schemas.CategoryOut])
def get_categories(domain: str = Query(...), db: Session = Depends(get_db)):
    dom = db.query(models.Domain).filter(models.Domain.code == domain).first()
    if not dom:
        raise HTTPException(404, "도메인 없음")
    return db.query(models.Category).filter(
        models.Category.domain_id == dom.id
    ).order_by(models.Category.sort_order).all()

@app.post("/api/categories", response_model=schemas.CategoryOut)
def create_category(data: schemas.CategoryCreate, db: Session = Depends(get_db),
                    admin: models.User = Depends(require_admin)):
    cat = models.Category(**data.model_dump())
    db.add(cat); db.commit(); db.refresh(cat)
    return cat

@app.put("/api/categories/{cat_id}", response_model=schemas.CategoryOut)
def update_category(cat_id: int, data: schemas.CategoryCreate,
                    db: Session = Depends(get_db),
                    admin: models.User = Depends(require_admin)):
    cat = db.query(models.Category).get(cat_id)
    if not cat:
        raise HTTPException(404)
    for k, v in data.model_dump().items():
        setattr(cat, k, v)
    db.commit(); db.refresh(cat)
    return cat

@app.delete("/api/categories/{cat_id}")
def delete_category(cat_id: int, db: Session = Depends(get_db),
                    admin: models.User = Depends(require_admin)):
    cat = db.query(models.Category).get(cat_id)
    if not cat:
        raise HTTPException(404)
    db.delete(cat); db.commit()
    return {"ok": True}

# ─────────────────────────────────────────
# ITEMS
# ─────────────────────────────────────────
def _item_with_price(item: models.Item) -> dict:
    d = {
        "id": item.id,
        "category_id": item.category_id,
        "name": item.name,
        "spec": item.spec,
        "unit": item.unit,
        "unit_weight": item.unit_weight,
        "is_weight_based": item.is_weight_based,
        "supplier": item.supplier,
        "is_active": item.is_active,
        "sort_order": item.sort_order,
        "category_name": item.category.name if item.category else "",
        "current_price": item.prices[0].price if item.prices else 0,
    }
    return d

@app.get("/api/items")
def get_items(domain: str = Query(...), category_id: Optional[int] = None,
              active_only: bool = True, db: Session = Depends(get_db)):
    dom = db.query(models.Domain).filter(models.Domain.code == domain).first()
    if not dom:
        raise HTTPException(404)
    q = (db.query(models.Item)
         .join(models.Category)
         .filter(models.Category.domain_id == dom.id))
    if category_id:
        q = q.filter(models.Item.category_id == category_id)
    if active_only:
        q = q.filter(models.Item.is_active == True)
    items = q.order_by(models.Category.sort_order, models.Item.sort_order).all()
    return [_item_with_price(i) for i in items]

@app.get("/api/items/{item_id}")
def get_item(item_id: int, db: Session = Depends(get_db)):
    item = db.query(models.Item).get(item_id)
    if not item:
        raise HTTPException(404)
    return _item_with_price(item)

@app.post("/api/items")
def create_item(data: schemas.ItemCreate, db: Session = Depends(get_db),
                admin: models.User = Depends(require_admin)):
    item_data = data.model_dump()
    initial_price = item_data.pop("initial_price", 0)
    item = models.Item(**item_data)
    db.add(item); db.flush()
    price = models.PriceMaster(item_id=item.id, price=initial_price)
    db.add(price)
    db.commit(); db.refresh(item)
    return _item_with_price(item)

@app.put("/api/items/{item_id}")
def update_item(item_id: int, data: schemas.ItemUpdate,
                db: Session = Depends(get_db),
                admin: models.User = Depends(require_admin)):
    item = db.query(models.Item).get(item_id)
    if not item:
        raise HTTPException(404)
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(item, k, v)
    db.commit(); db.refresh(item)
    return _item_with_price(item)

@app.delete("/api/items/{item_id}")
def delete_item(item_id: int, db: Session = Depends(get_db),
                admin: models.User = Depends(require_admin)):
    item = db.query(models.Item).get(item_id)
    if not item:
        raise HTTPException(404)
    item.is_active = False
    db.commit()
    return {"ok": True}

# ─────────────────────────────────────────
# PRICES
# ─────────────────────────────────────────
@app.get("/api/prices/{item_id}/history", response_model=List[schemas.PriceHistoryOut])
def price_history(item_id: int, db: Session = Depends(get_db)):
    return (db.query(models.PriceHistory)
            .filter(models.PriceHistory.item_id == item_id)
            .order_by(models.PriceHistory.changed_at.desc()).all())

@app.put("/api/prices/{item_id}")
def update_price(item_id: int, data: schemas.PriceUpdate,
                 db: Session = Depends(get_db),
                 admin: models.User = Depends(require_admin)):
    item = db.query(models.Item).get(item_id)
    if not item:
        raise HTTPException(404)
    old_price = item.prices[0].price if item.prices else 0
    # 이력 저장
    hist = models.PriceHistory(
        item_id=item_id, old_price=old_price, new_price=data.price,
        changed_by=admin.username, note=data.note
    )
    db.add(hist)
    # 새 단가 마스터
    pm = models.PriceMaster(item_id=item_id, price=data.price)
    db.add(pm)
    db.commit()
    return {"ok": True, "old_price": old_price, "new_price": data.price}

@app.post("/api/prices/bulk-update")
def bulk_update_prices(
    domain: str = Query(...),
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin)
):
    """CSV 일괄 업로드용 엔드포인트 (프론트에서 JSON 배열로 전달)"""
    pass

@app.post("/api/prices/bulk-json")
def bulk_update_prices_json(
    updates: List[dict],
    db: Session = Depends(get_db),
    admin: models.User = Depends(require_admin)
):
    """[{item_id, price, note}] 배열로 일괄 단가 업데이트"""
    for u in updates:
        item = db.query(models.Item).get(u["item_id"])
        if not item:
            continue
        old = item.prices[0].price if item.prices else 0
        db.add(models.PriceHistory(
            item_id=u["item_id"], old_price=old, new_price=u["price"],
            changed_by=admin.username, note=u.get("note", "일괄 업데이트")
        ))
        db.add(models.PriceMaster(item_id=u["item_id"], price=u["price"]))
    db.commit()
    return {"ok": True, "updated": len(updates)}

# ─────────────────────────────────────────
# PERCENTAGES
# ─────────────────────────────────────────
@app.get("/api/percentages", response_model=List[schemas.PercentOut])
def get_percentages(domain: str = Query(...), db: Session = Depends(get_db)):
    dom = db.query(models.Domain).filter(models.Domain.code == domain).first()
    if not dom:
        raise HTTPException(404)
    return (db.query(models.PercentMaster)
            .filter(models.PercentMaster.domain_id == dom.id)
            .order_by(models.PercentMaster.sort_order).all())

@app.put("/api/percentages/{pct_id}", response_model=schemas.PercentOut)
def update_percentage(pct_id: int, data: schemas.PercentUpdate,
                      db: Session = Depends(get_db),
                      admin: models.User = Depends(require_admin)):
    pct = db.query(models.PercentMaster).get(pct_id)
    if not pct:
        raise HTTPException(404)
    pct.percent = data.percent
    db.commit(); db.refresh(pct)
    return pct

# ─────────────────────────────────────────
# QUOTES
# ─────────────────────────────────────────
@app.get("/api/quotes")
def list_quotes(domain: str = Query(...), db: Session = Depends(get_db),
                user: models.User = Depends(require_user)):
    dom = db.query(models.Domain).filter(models.Domain.code == domain).first()
    if not dom:
        raise HTTPException(404)
    quotes = (db.query(models.Quote)
              .filter(models.Quote.domain_id == dom.id)
              .order_by(models.Quote.created_at.desc()).all())
    result = []
    for q in quotes:
        mat = sum(i.amount for i in q.items)
        result.append({
            "id": q.id, "quote_number": q.quote_number,
            "field_name": q.field_name, "client_name": q.client_name,
            "work_name": q.work_name, "quote_date": q.quote_date,
            "status": q.status, "created_at": q.created_at,
            "material_total": mat
        })
    return result

@app.get("/api/quotes/{quote_id}", response_model=schemas.QuoteOut)
def get_quote(quote_id: int, db: Session = Depends(get_db),
              user: models.User = Depends(require_user)):
    q = db.query(models.Quote).get(quote_id)
    if not q:
        raise HTTPException(404)
    return q

@app.post("/api/quotes", response_model=schemas.QuoteOut)
def create_quote(data: schemas.QuoteCreate, db: Session = Depends(get_db),
                 user: models.User = Depends(require_user)):
    # 견적번호 자동 생성
    count = db.query(models.Quote).count() + 1
    qnum = f"EST-{datetime.now().strftime('%Y%m')}-{count:04d}"
    q = models.Quote(
        domain_id=data.domain_id,
        quote_number=qnum,
        field_name=data.field_name,
        client_name=data.client_name,
        work_name=data.work_name,
        quote_date=data.quote_date,
        tax_included=data.tax_included,
        margin_rate=data.margin_rate,
        labor_pct=data.labor_pct,
        expense_pct=data.expense_pct,
        management_pct=data.management_pct,
        profit_pct=data.profit_pct,
        tax_pct=data.tax_pct,
        note=data.note,
        created_by=user.username,
    )
    db.add(q); db.flush()
    for idx, item in enumerate(data.items):
        qi = models.QuoteItem(
            quote_id=q.id, sort_order=idx,
            **item.model_dump()
        )
        db.add(qi)
    db.commit(); db.refresh(q)
    return q

@app.put("/api/quotes/{quote_id}", response_model=schemas.QuoteOut)
def update_quote(quote_id: int, data: schemas.QuoteCreate,
                 db: Session = Depends(get_db),
                 user: models.User = Depends(require_user)):
    q = db.query(models.Quote).get(quote_id)
    if not q:
        raise HTTPException(404)
    for k in ["field_name","client_name","work_name","quote_date",
              "tax_included","margin_rate","labor_pct","expense_pct",
              "management_pct","profit_pct","tax_pct","note"]:
        setattr(q, k, getattr(data, k))
    # 기존 아이템 삭제 후 재삽입
    db.query(models.QuoteItem).filter(models.QuoteItem.quote_id == quote_id).delete()
    for idx, item in enumerate(data.items):
        qi = models.QuoteItem(quote_id=q.id, sort_order=idx, **item.model_dump())
        db.add(qi)
    db.commit(); db.refresh(q)
    return q

@app.delete("/api/quotes/{quote_id}")
def delete_quote(quote_id: int, db: Session = Depends(get_db),
                 user: models.User = Depends(require_user)):
    q = db.query(models.Quote).get(quote_id)
    if not q:
        raise HTTPException(404)
    db.query(models.QuoteItem).filter(models.QuoteItem.quote_id == quote_id).delete()
    db.delete(q); db.commit()
    return {"ok": True}

# ─────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────
@app.get("/api/quotes/{quote_id}/excel")
def export_excel(quote_id: int, db: Session = Depends(get_db),
                 user: models.User = Depends(require_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    q = db.query(models.Quote).get(quote_id)
    if not q:
        raise HTTPException(404)

    wb = Workbook()
    ws = wb.active
    ws.title = "견적서"

    # 스타일 정의
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=16)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    subheader_fill = PatternFill("solid", fgColor="2E75B6")
    cat_fill = PatternFill("solid", fgColor="BDD7EE")
    white_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center")

    # 제목
    ws.merge_cells("A1:J1")
    domain_name = q.domain.name if q.domain else ""
    ws["A1"] = f"{domain_name} 견적서"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    # 기본 정보
    ws["A3"] = "현장명"; ws["B3"] = q.field_name
    ws["A4"] = "거래처명"; ws["B4"] = q.client_name
    ws["A5"] = "공사명"; ws["B5"] = q.work_name
    ws["G3"] = "견적번호"; ws["H3"] = q.quote_number
    ws["G4"] = "견적일자"; ws["H4"] = q.quote_date
    for r in [3, 4, 5]:
        ws.cell(r, 1).font = header_font

    # 헤더
    row = 7
    headers = ["번호", "공종", "품목명", "규격", "단위", "수량", "단가", "금액", "비고"]
    col_widths = [6, 15, 25, 15, 8, 10, 12, 15, 15]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row, ci, h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    row += 1

    # 아이템
    mat_total = 0
    current_cat = None
    num = 1
    from openpyxl.styles import numbers as xl_numbers
    for qi in q.items:
        if qi.category_name != current_cat:
            current_cat = qi.category_name
            ws.merge_cells(f"A{row}:I{row}")
            cell = ws.cell(row, 1, f"【 {current_cat} 】")
            cell.font = Font(bold=True, size=10)
            cell.fill = cat_fill
            cell.alignment = center
            for ci in range(1, 10):
                ws.cell(row, ci).border = border
            row += 1

        ws.cell(row, 1, num).alignment = center
        ws.cell(row, 2, qi.category_name)
        ws.cell(row, 3, qi.item_name)
        ws.cell(row, 4, qi.spec)
        ws.cell(row, 5, qi.unit).alignment = center
        ws.cell(row, 6, qi.quantity).number_format = "#,##0.00"
        ws.cell(row, 7, qi.unit_price).number_format = "#,##0"
        ws.cell(row, 8, qi.amount).number_format = "#,##0"
        ws.cell(row, 9, qi.note)
        for ci in range(1, 10):
            ws.cell(row, ci).border = border
        mat_total += qi.amount
        num += 1
        row += 1

    # 합계
    row += 1
    labor = mat_total * q.labor_pct / 100
    expense = mat_total * q.expense_pct / 100
    mgmt = mat_total * q.management_pct / 100
    profit = mat_total * q.profit_pct / 100
    subtotal = mat_total + labor + expense + mgmt + profit
    tax = subtotal * q.tax_pct / 100
    total = subtotal + tax

    summary = [
        ("자재비 합계", mat_total),
        (f"노무비 ({q.labor_pct}%)", labor),
        (f"경비 ({q.expense_pct}%)", expense),
        (f"일반관리비 ({q.management_pct}%)", mgmt),
        (f"이윤 ({q.profit_pct}%)", profit),
        ("공급가액", subtotal),
        (f"부가세 ({q.tax_pct}%)", tax),
        ("합계", total),
    ]
    for label, val in summary:
        ws.merge_cells(f"A{row}:G{row}")
        lc = ws.cell(row, 1, label)
        lc.font = Font(bold=True)
        lc.alignment = Alignment(horizontal="right")
        vc = ws.cell(row, 8, val)
        vc.number_format = "#,##0"
        vc.font = Font(bold=True)
        for ci in range(1, 10):
            ws.cell(row, ci).border = border
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"견적서_{q.quote_number}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"}
    )
